import openpyxl
import os
from io import BytesIO

from django.conf import settings
from django.core.files.base import ContentFile

from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch, mm


class ReportsFileGenerator:
    def __init__(self, data_entries, headers, context):
        # Initialize Headers, and data
        self.headers = headers
        self.data_entries = data_entries
        self.templates_path = os.path.join(
            settings.BASE_DIR, "data_tables", "file_generation", "templates"
        )
        self.context = context
        # Get the project and category name
        if not self.context["project"]:
            self.context["project"] = "Global Statistics"
        else:
            self.context["project"] = self.context["project"].name
        if not self.context["category"]:
            self.context["category"] = "Global Statistics"
        else:
            self.context["category"] = self.context["category"].name

    def generate_excel_file(self):
        template_path = os.path.join(self.templates_path, "template.xlsx")
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        starting_row = 6
        current_row = starting_row
        date_without_tz = self.context["date"].replace(tzinfo=None)
        ws["B2"] = self.context["project"]
        ws["B3"] = self.context["category"]
        ws["B4"] = date_without_tz.strftime("%Y-%m-%d %H:%M:%S")

        for row_num, header in enumerate(self.headers, start=1):
            cell = ws.cell(row=current_row, column=row_num, value=header)

        current_row += 1

        for entry in self.data_entries:
            row = []
            for field in self.headers:
                value = entry.get(field)
                if value is not None:
                    row.append(value)
                else:
                    try:
                        row.append(getattr(entry.instance, field))
                    except AttributeError:
                        row.append(None)
            for col_num, data in enumerate(row, start=1):
                ws.cell(row=current_row, column=col_num, value=data)
            current_row += 1

        output = BytesIO()
        wb.save(output)
        file = ContentFile(output.getvalue(), self.context["name"] + ".xlsx")
        return file

    @property
    def generate_pdf_file(self):
        headers = []
        for header in self.headers:
            if "_" in header:
                if "PWD" in header:
                    short_header = "f_PWD" if "female" in header else "m_PWD"
                elif "female" in header:
                    short_header = "f_" + "_".join(header.split("_")[1:])
                elif "name" in header:
                    short_header = header
                else:
                    short_header = "m_" + "_".join(header.split("_")[1:])
                headers.append(short_header)
            else:
                headers.append(header)

        output = BytesIO()

        document = SimpleDocTemplate(output, pagesize=landscape(letter))
        logo_path = os.path.join(self.templates_path, "logo.png")
        im = Image(logo_path, 4 * inch, 1 * inch, hAlign="LEFT")

        elements = [im, Spacer(1, 10 * mm)]
        # Data for table creation
        data = [headers]
        for entry in self.data_entries:
            row = []
            for field in self.headers:
                value = entry.get(field)
                if value is not None:
                    row.append(value)
                else:
                    try:
                        row.append(getattr(entry.instance, field))
                    except AttributeError:
                        row.append(None)
            data.append(row)

        header_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.gray),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )

        row_style = TableStyle([("GRID", (0, 0), (-1, -1), 1, colors.black)])

        hAlign = "LEFT" if len(headers) <= 15 else "CENTER"

        table = Table(data, repeatRows=1, hAlign=hAlign)

        table.setStyle(header_style)
        table.setStyle(row_style)

        elements.append(table)

        # Build PDF
        document.build(elements)

        output.seek(0)
        file = ContentFile(output.getvalue(), self.context["name"] + ".pdf")
        return file


if __name__ == "__main__":
    print("main")
