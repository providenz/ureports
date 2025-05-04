import os
import django
import random
from faker import Faker
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "volunteer_reports.settings")
django.setup()

from accounts.models import CustomUser
from reports.models import Project
from data_tables.models import DataTable

fake = Faker()


def add_data_from_excel(excel_file_path):
    print("Загрузка файла Excel...")
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    donors = CustomUser.objects.filter(is_superuser=False)
    print(f"Найдено {len(donors)} доноров.")

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        (
            category_people_str,
            gender,
            age,
            oblast,
            rayon,
            hromada,
            place,
            address,
            female_0_17,
            male_0_17,
            female_18_59,
            male_18_59,
            female_60plus,
            male_60plus,
        ) = row

        donor = random.choice(donors)
        donor_projects = donor.donated_projects.all()

        if donor_projects:
            print(f"Обработка строки {idx}...")
            chosen_project = random.choice(donor_projects)

            data_instance = DataTable(
                category_name=category_people_str,
                gender=gender,
                age=age,
                oblast=oblast,
                rayon=rayon,
                hromada=hromada,
                place=place,
                address=address,
                female_0_17=female_0_17,
                male_0_17=male_0_17,
                female_18_59=female_18_59,
                male_18_59=male_18_59,
                female_60plus=female_60plus,
                male_60plus=male_60plus,
                user=donor,
                project=chosen_project,
                category=random.choice(chosen_project.categories.all())
                if chosen_project.categories.exists()
                else None,
            )
            data_instance.save()
            print(f"Данные из строки {idx} добавлены.")
        else:
            print(f"У донора {donor} нет проектов. Строка {idx} пропущена.")


if __name__ == "__main__":
    excel_file_path = "test_data.xlsx"
    print("Запуск скрипта...")
    add_data_from_excel(excel_file_path)
    print("Работа скрипта завершена.")
